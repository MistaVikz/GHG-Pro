import sys
import os
import pandas as pd
import numpy as np
import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Side, PatternFill
from tabulate import tabulate

def load_and_process_data(input_file: str = 'GHG_Data.xlsx') -> tuple[int, int, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Load and process data from Excel file.

    Parameters:
    input_file (str): The path to the input Excel file. Default is 'GHG_Data.xlsx'.

    Returns:
    num_buckets (int): The number of risk buckets.
    num_risk_factors (int): The number of risk factors.
    df_project (pd.DataFrame): A DataFrame containing project data.
    df_default_rates (pd.DataFrame): A DataFrame containing default rates data.
    df_recovery_potential (pd.DataFrame): A DataFrame containing recovery potential data.
    df_model (pd.DataFrame): A DataFrame containing model configuration data.

    Notes:
    This function assumes that the input Excel file contains four sheets named 'Project Data', 'Default Rates', 'Recovery Potential', and 'Model Config'.
    """
    # Load Data
    df_project = pd.read_excel(input_file, sheet_name='Project Data')  
    df_default_rates = pd.read_excel(input_file, sheet_name='Default Rates')
    df_recovery_potential = pd.read_excel(input_file, sheet_name='Recovery Potential')
    df_model = pd.read_excel(input_file, sheet_name='Model Config')

    # Set index for default rates and recovery potential dataframes
    df_default_rates = df_default_rates.set_index('Unnamed: 0')
    df_recovery_potential = df_recovery_potential.set_index('Unnamed: 0')

    # Convert Data Types
    df_default_rates.columns = df_default_rates.columns.astype(int)
    df_recovery_potential.columns = df_recovery_potential.columns.astype(int)

    # Replace NaN with 0 in risk bucket factors and weights columns
    risk_columns = [col for col in df_project.columns if 'risk_bucket' in col]
    df_project[risk_columns] = df_project[risk_columns].fillna(0)

    # Convert 'Screening Date' to datetime
    df_project['screening_date'] = pd.to_datetime(df_project['screening_date'])

    # Calculate the number of risk buckets
    num_buckets = max(int(col.split('_')[2]) for col in df_project.columns if 'risk_bucket' in col)

    # Calculate the number of risk factors
    num_risk_factors = len([col for col in df_project.columns if 'factor' in col and f"risk_bucket_1_factor" in col])

    return num_buckets, num_risk_factors, df_project, df_default_rates, df_recovery_potential, df_model

def create_output_folder() -> str:
    """
    Create a time-stamped folder for the simulation output and summary output
    """
    # Get the current date and time
    now = datetime.datetime.now()
    date_time = now.strftime("%Y%m%d_%H%M%S")

    # Get the current directory of the script
    current_dir = os.path.dirname(__file__)

    # Construct the path to the output directory
    output_dir = os.path.join(current_dir, '..', 'output')

    # Create a new folder with the timestamp
    folder_name = os.path.join(output_dir, date_time)
    os.makedirs(folder_name, exist_ok=True)

    return folder_name

def export_ghg_data(output_folder: str, df_project: 'pd.DataFrame', df_default_rates: 'pd.DataFrame', 
                    df_recovery_potential: 'pd.DataFrame', df_model: 'pd.DataFrame') -> None:
    """
    Export GHG data to an Excel file.

    Parameters:
    output_folder (str): The folder to save the Excel file in.
    df_project (pd.DataFrame): A DataFrame containing project data.
    df_default_rates (pd.DataFrame): A DataFrame containing default rates data.
    df_recovery_potential (pd.DataFrame): A DataFrame containing recovery potential data.
    df_model (pd.DataFrame): A DataFrame containing model configuration data.

    Notes:
    This function exports the input DataFrames to four separate worksheets in the 'GHG_Data.xlsx' file.
    The 'Default Rates' and 'Recovery Potential' worksheets include "Investment", "Speculative", and "C" next to the values.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Project Data'

    for row in dataframe_to_rows(df_project, index=False):
        ws.append(row)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 25
    ws.column_dimensions['B'].width = 35

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.row % 2 == 0:
                cell.fill = PatternFill(start_color='C5C5C5', fill_type='solid')

    for cell in ws["1:1"]:
        cell.font = Font(bold=True, underline='single')

    wb.create_sheet('Default Rates')
    ws = wb['Default Rates']
    ws.append([""] + list(df_default_rates.columns))
    ws.append(["Investment"] + list(df_default_rates.loc["Investment"]))
    ws.append(["Speculative"] + list(df_default_rates.loc["Speculative"]))
    ws.append(["C"] + list(df_default_rates.loc["C"]))

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 15
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.row % 2 == 0:
                cell.fill = PatternFill(start_color='C5C5C5', fill_type='solid')

    for cell in ws["1:1"]:
        cell.font = Font(bold=True, underline='single')

    wb.create_sheet('Recovery Potential')
    ws = wb['Recovery Potential']
    ws.append([""] + list(df_recovery_potential.columns))
    ws.append(["Investment"] + list(df_recovery_potential.loc["Investment"]))
    ws.append(["Speculative"] + list(df_recovery_potential.loc["Speculative"]))
    ws.append(["C"] + list(df_recovery_potential.loc["C"]))

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 15
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.row % 2 == 0:
                cell.fill = PatternFill(start_color='C5C5C5', fill_type='solid')

    for cell in ws["1:1"]:
        cell.font = Font(bold=True, underline='single')

    wb.create_sheet('Model Config')
    ws = wb['Model Config']

    for row in dataframe_to_rows(df_model, index=False):
        ws.append(row)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 25

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.row % 2 == 0:
                cell.fill = PatternFill(start_color='C5C5C5', fill_type='solid')

    for cell in ws["1:1"]:
        cell.font = Font(bold=True, underline='single')

    wb.save(f"{output_folder}/GHG_Data_Simulation.xlsx")

def export_project_risk_output(output_folder: str, top_projects: 'pd.DataFrame', bottom_projects: 'pd.DataFrame', 
                               country_table: 'pd.DataFrame', technology_table: 'pd.DataFrame', 
                               counterparty_table: 'pd.DataFrame', df_counts: 'pd.DataFrame', 
                               total_volumes_per_year: 'pd.DataFrame') -> None:
    """
    Exports project risk output to an Excel file.

    Parameters:
    output_folder (str): The folder to save the Excel file in.
    top_projects (pd.DataFrame): Top performing projects.
    bottom_projects (pd.DataFrame): Bottom performing projects.
    country_table (pd.DataFrame): Country table.
    technology_table (pd.DataFrame): Technology table.
    counterparty_table (pd.DataFrame): Counterparty table.
    df_counts (pd.DataFrame): Project rating distribution counts.
    total_volumes_per_year (pd.DataFrame): Total volumes per year data.

    Returns:
    None

    Notes:
    This function creates an Excel file named 'project_risk_output.xlsx' and writes the input DataFrames to separate worksheets in the file. The worksheets are ordered as follows: Project Rating Distribution, Highest Performing Projects, Lowest Performing Projects, Annual Project Volumes, Overall Project Rating Distribution by Country, Overall Project Rating Distribution by Technology, Overall Project Rating Distribution by Counterparty.
    """
    
    wb = Workbook()
    ws_df_counts = wb.active
    ws_df_counts.title = "Rating Distribution"
    ws_top_projects = wb.create_sheet("Highest Performing")
    ws_bottom_projects = wb.create_sheet("Lowest Performing")
    ws_total_volumes_per_year = wb.create_sheet("Annual Volumes")
    ws_country_table = wb.create_sheet("Country Distribution")
    ws_technology_table = wb.create_sheet("Technology Distribution")
    ws_counterparty_table = wb.create_sheet("Counterparty Distribution")

    wb._sheets = [ws_df_counts, ws_top_projects, ws_bottom_projects, ws_total_volumes_per_year, ws_country_table, ws_technology_table, ws_counterparty_table]

    ws_df_counts['A1'] = "Project Rating Distribution"
    ws_df_counts['A1'].font = Font(bold=True, size=16)
    ws_top_projects['A1'] = "Highest Performing Projects"
    ws_top_projects['A1'].font = Font(bold=True, size=16)
    ws_bottom_projects['A1'] = "Lowest Performing Projects"
    ws_bottom_projects['A1'].font = Font(bold=True, size=16)
    ws_total_volumes_per_year['A1'] = "Annual Project Volumes"
    ws_total_volumes_per_year['A1'].font = Font(bold=True, size=16)
    ws_country_table['A1'] = "Overall Project Rating Distribution by Country"
    ws_country_table['A1'].font = Font(bold=True, size=16)
    ws_technology_table['A1'] = "Overall Project Rating Distribution by Technology"
    ws_technology_table['A1'].font = Font(bold=True, size=16)
    ws_counterparty_table['A1'] = "Overall Project Rating Distribution by Counterparty"
    ws_counterparty_table['A1'].font = Font(bold=True, size=16)

    for row in dataframe_to_rows(df_counts, index=False, header=True):
        ws_df_counts.append(row)
    for row in dataframe_to_rows(top_projects, index=False, header=True):
        ws_top_projects.append(row)
    for row in dataframe_to_rows(bottom_projects, index=False, header=True):
        ws_bottom_projects.append(row)
    for row in dataframe_to_rows(total_volumes_per_year, index=False, header=True):
        ws_total_volumes_per_year.append(row)

    rows = dataframe_to_rows(country_table, index=False, header=True)
    header = next(rows)
    ws_country_table.append(header)
    for row in rows:
        if row != [''] * len(row):
            ws_country_table.append(row)

    rows = dataframe_to_rows(technology_table, index=False, header=True)
    header = next(rows)
    ws_technology_table.append(header)
    for row in rows:
        if row != [''] * len(row):
            ws_technology_table.append(row)

    rows = dataframe_to_rows(counterparty_table, index=False, header=True)
    header = next(rows)
    ws_counterparty_table.append(header)
    for row in rows:
        if row != [''] * len(row):
            ws_counterparty_table.append(row)

    for ws in [ws_top_projects, ws_bottom_projects, ws_country_table, ws_technology_table, ws_counterparty_table, ws_df_counts, ws_total_volumes_per_year]:
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20
    ws_top_projects.column_dimensions['C'].width = 35
    ws_bottom_projects.column_dimensions['C'].width = 35
    ws_top_projects.column_dimensions['H'].width = 26
    ws_bottom_projects.column_dimensions['H'].width = 26
    ws_total_volumes_per_year.column_dimensions['C'].width = 25
    ws_df_counts.column_dimensions['A'].width = 26

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for ws in [ws_top_projects, ws_bottom_projects, ws_country_table, ws_technology_table, ws_counterparty_table, ws_df_counts, ws_total_volumes_per_year]:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border

    for ws in [ws_top_projects, ws_bottom_projects, ws_country_table, ws_technology_table, ws_counterparty_table, ws_df_counts, ws_total_volumes_per_year]:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.row % 2 == 0:
                    cell.fill = PatternFill(start_color='C5C5C5', fill_type='solid')

    for ws in [ws_top_projects, ws_bottom_projects, ws_country_table, ws_technology_table, ws_counterparty_table, ws_df_counts, ws_total_volumes_per_year]:
        for cell in ws["2:2"]:
            cell.font = Font(bold=True, underline='single')

    wb.save(f"{output_folder}/Project_Risk_Summary_Data.xlsx")

def valid_project_data(df: pd.DataFrame, num_buckets: int, num_factors: int) -> bool:
    """
    Validates the data in the project DataFrame.

    Parameters:
    df (pd.DataFrame): The project DataFrame to validate.
    num_buckets (int): The number of risk buckets.
    num_factors (int): The number of risk factors.

    Returns:
    bool: True if the data is valid, False otherwise.
    """

    required_columns = ['project_id', 'project_name', 'contract_duration', 'country', 'technology', 'counterparty', 'start_year', 'screening_date']
    for year in range(1, 11):
        required_columns.append(f'offered_volume_year_{year}')
    for bucket in range(1, num_buckets + 1):
        for factor in range(1, num_factors + 1):
            required_columns.append(f'risk_bucket_{bucket}_factor_{factor}')
            required_columns.append(f'risk_bucket_{bucket}_weight_{factor}')

    if not all(column in df.columns for column in required_columns):
        print("Error: Not all required columns are present.")
        return False
    
    if df['project_id'].isnull().any() or df['project_id'].duplicated().any():
        print("Error: Project ID column contains null or duplicate values.")
        return False
    
    if df['contract_duration'].dtype != 'int64' or (df['contract_duration'] < 1).any() or (df['contract_duration'] > 10).any():
        print("Error: Contract Duration column contains invalid values. Values must be integers between 1 and 10.")
        return False
    
    if df['start_year'].dtype != 'int64' or (df['start_year'] < 2000).any():
        print("Error: Start Year column contains invalid values. Values must be integers greater or equal to 2000.")
        return False
        
    if df['screening_date'].dtype != 'datetime64[ns]':
        print("Error: Screening Date column contains invalid values. Values must be dates.")
        return False
    
    for column in ['project_name', 'technology', 'country', 'counterparty']:
        if df[column].dtype != 'object':
            print(f"Error: {column} column contains invalid values. Values must be strings.")
            return False

    for bucket in range(1, num_buckets + 1):
        weight_columns = [f'risk_bucket_{bucket}_weight_{factor}' for factor in range(1, num_factors + 1)]
        if not np.all(np.round(df[weight_columns].sum(axis=1), 6) == 1):
            print(f"Error: Weights for risk bucket {bucket} must sum to 1.")
            return False
    
    return True

def display_project_risk_output(df_counts: pd.DataFrame, top_projects: pd.DataFrame, bottom_projects: pd.DataFrame, 
                                country_table: pd.DataFrame, technology_table: pd.DataFrame, counterparty_table: pd.DataFrame, 
                                total_volumes_per_year: pd.DataFrame) -> None:
    """
    Prints project risk output dataframes in a tabular format.

    Parameters:
    df_counts (pd.DataFrame): Project rating distribution dataframe.
    top_projects (pd.DataFrame): Top projects dataframe.
    bottom_projects (pd.DataFrame): Bottom projects dataframe.
    country_table (pd.DataFrame): Country table dataframe.
    technology_table (pd.DataFrame): Technology table dataframe.
    counterparty_table (pd.DataFrame): Counterparty table dataframe.
    total_volumes_per_year (pd.DataFrame): Total volumes per year dataframe.
    """
    
    # Define the table format constant
    TABLE_FORMAT = 'fancy_grid'

    # Don't print if cmd line argument is set
    if len(sys.argv) > 1 and sys.argv[1] == '--noprint':
        return

    print()
    print("Project Rating Distribution")
    print(tabulate(df_counts, headers='keys', tablefmt=TABLE_FORMAT, showindex=False))
    print()

    print("Highest Performing Projects")
    print(tabulate(top_projects, headers='keys', tablefmt=TABLE_FORMAT, showindex=False))
    print()

    print("Lowest Performing Projects")
    print(tabulate(bottom_projects, headers='keys', tablefmt=TABLE_FORMAT, showindex=False))
    print()

    print("Overall Project Rating Distribution by Country")
    print(tabulate(country_table, headers='keys', tablefmt=TABLE_FORMAT, showindex=False))
    print()

    print("Overall Project Rating Distribution by Technology")
    print(tabulate(technology_table, headers='keys', tablefmt=TABLE_FORMAT, showindex=False))
    print()

    print("Overall Project Rating Distribution by Counterparty")
    print(tabulate(counterparty_table, headers='keys', tablefmt=TABLE_FORMAT, showindex=False))
    print()

    print("Annual Project Volumes")
    print(tabulate(total_volumes_per_year, headers='keys', tablefmt=TABLE_FORMAT, showindex=False))

def check_df_format(df_default_rates: pd.DataFrame, df_recovery_potential: pd.DataFrame) -> bool:
    """
    Checks if two dataframes are formatted correctly.

    "Formatted correctly" means that the dataframes should have the same shape, 
    the same column names, the same index values, and the correct row names.

    Parameters:
    df_default_rates (pd.DataFrame): The first dataframe to compare.
    df_recovery_potential (pd.DataFrame): The second dataframe to compare.

    Returns:
    bool: True if the dataframes are formatted correctly, False otherwise.
    """

    # Define the expected row names and column names
    expected_index = ['Investment', 'Speculative', 'C']
    expected_columns = list(range(1, 11))

    # Check if both dataframes have the same shape
    if df_default_rates.shape != df_recovery_potential.shape:
        print("Error: Dataframes do not have the same shape.")
        print(f"df_default_rates shape: {df_default_rates.shape}")
        print(f"df_recovery_potential shape: {df_recovery_potential.shape}")
        return False

    # Check if both dataframes have the correct row names and column names
    if (list(df_default_rates.index) != expected_index or 
        list(df_recovery_potential.index) != expected_index or 
        list(df_default_rates.columns) != expected_columns or 
        list(df_recovery_potential.columns) != expected_columns):
        print("Error: Dataframes do not have the correct row names or column names.")
        print(f"df_default_rates index: {list(df_default_rates.index)}")
        print(f"df_recovery_potential index: {list(df_recovery_potential.index)}")
        print(f"df_default_rates columns: {list(df_default_rates.columns)}")
        print(f"df_recovery_potential columns: {list(df_recovery_potential.columns)}")
        return False

    # Check if all values in both dataframes are numbers greater or equal to zero
    if not ((df_default_rates.apply(lambda x: x >= 0) & df_default_rates.apply(lambda x: pd.to_numeric(x, errors='coerce')).notnull()).all().all() and
            (df_recovery_potential.apply(lambda x: x >= 0) & df_recovery_potential.apply(lambda x: pd.to_numeric(x, errors='coerce')).notnull()).all().all()):
        print("Error: Dataframes contain invalid values.")
        return False

    return True

def valid_model(df_model: pd.DataFrame, risk_bucket_count: int, risk_factor_count: int) -> bool:
    """
    Validate the model configuration data.

    Parameters:
    df_model (pd.DataFrame): DataFrame containing the model configuration data.
    risk_bucket_count (int): Number of risk buckets in the model.
    risk_factor_count (int): Number of risk factors in the model.

    Returns:
    bool: True if the model configuration data is valid, False otherwise.
    """

    # Calculate the required columns
    required_columns = ['model_id', 'model_name', 'num_buckets', 'num_factors', 'last_saved']
    required_columns.extend(f'risk_bucket_{i}_name' for i in range(1, risk_bucket_count + 1))
    required_columns.extend(f'risk_bucket_{i}_factor_{j}_name' for i in range(1, risk_bucket_count + 1) for j in range(1, risk_factor_count + 1))
    required_columns.extend(f'risk_bucket_{i}_factor_{j}_rules' for i in range(1, risk_bucket_count + 1) for j in range(1, risk_factor_count + 1))

    # Check if all required columns are present
    if set(required_columns) != set(df_model.columns):
        print("Error: Not all required columns are present.")
        return False

    # Check if there is only one row in the dataframe
    if df_model.shape[0] != 1:
        print("Error: There should be only one row in the dataframe.")
        return False

    return True